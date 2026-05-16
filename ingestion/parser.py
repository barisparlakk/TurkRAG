"""Document text extraction for PDF, DOCX, TXT, XLSX, and CSV files."""

import logging
import re
from pathlib import Path

logger = logging.getLogger(__name__)


def parse_document(file_path: str) -> str:
    """Extract raw text from a PDF, DOCX, TXT, XLSX, or CSV file.

    Strips common header/footer noise heuristically.
    Returns clean plain text suitable for chunking.
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    logger.info("Parsing document: %s (type=%s)", path.name, suffix)

    if suffix == ".pdf":
        text = _parse_pdf(file_path)
    elif suffix == ".docx":
        text = _parse_docx(file_path)
    elif suffix == ".txt":
        text = _parse_txt(file_path)
    elif suffix in (".xlsx", ".xls"):
        text = _parse_excel(file_path)
    elif suffix == ".csv":
        text = _parse_csv(file_path)
    else:
        raise ValueError(f"Unsupported file type: {suffix}. Supported: .pdf, .docx, .txt, .xlsx, .csv")

    text = _clean_text(text)
    logger.info("Parsed %d characters from %s", len(text), path.name)
    return text


def _parse_pdf(file_path: str) -> str:
    """Extract text from a PDF using pdfplumber."""
    import pdfplumber

    pages = []
    with pdfplumber.open(file_path) as pdf:
        for i, page in enumerate(pdf.pages):
            page_text = page.extract_text()
            if page_text:
                pages.append(page_text)
                logger.debug("PDF page %d: %d chars", i + 1, len(page_text))
    return "\n\n".join(pages)


def _parse_docx(file_path: str) -> str:
    """Extract text from a DOCX file using python-docx."""
    from docx import Document

    doc = Document(file_path)
    paragraphs = []
    for para in doc.paragraphs:
        text = para.text.strip()
        if text:
            paragraphs.append(text)
    return "\n\n".join(paragraphs)


def _parse_txt(file_path: str) -> str:
    """Read a plain text file with UTF-8 encoding (fallback to latin-1)."""
    try:
        return Path(file_path).read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return Path(file_path).read_text(encoding="latin-1")


def _row_to_sentence(header: list, values: list) -> str:
    """Render header+value lists as 'Key: Value. Key: Value.' sentence."""
    pairs = [f"{h}: {v}" for h, v in zip(header, values) if v]
    return (". ".join(pairs) + ".") if pairs else ""


def _parse_excel(file_path: str) -> str:
    """Convert an Excel workbook to structured plain text.

    Each sheet becomes a section. The first non-empty row is treated as a
    header and subsequent rows are rendered as 'Sütun: Değer' pairs so the
    chunker can split on natural sentence boundaries.
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row (first row with at least one non-None cell)
        header: list = []
        data_rows: list = []
        for i, row in enumerate(rows):
            values = [str(v).strip() if v is not None else "" for v in row]
            if not any(values):
                continue
            if not header:
                header = values
            else:
                data_rows.append(values)

        if not header:
            continue

        section_lines = [f"=== Sayfa: {sheet_name} ==="]
        for row_vals in data_rows:
            if not any(v for v in row_vals):
                continue
            sentence = _row_to_sentence(header, row_vals)
            if sentence:
                section_lines.append(sentence)
        sections.append("\n".join(section_lines))

    result = "\n\n".join(sections)
    logger.info("Excel parsed: %d sheet(s), %d chars", len(wb.sheetnames), len(result))
    return result


def _parse_csv(file_path: str) -> str:
    """Convert a CSV file to structured plain text.

    Each row is rendered as 'Sütun: Değer' pairs, one per line.
    Handles UTF-8 and latin-1 encodings automatically.
    """
    import csv

    def _read(encoding: str) -> str:
        lines = []
        with open(file_path, newline="", encoding=encoding) as f:
            reader = csv.DictReader(f)
            header = reader.fieldnames or []
            for row in reader:
                sentence = _row_to_sentence(header, [row.get(h, "") for h in header])
                if sentence:
                    lines.append(sentence)
        return "\n".join(lines)

    try:
        text = _read("utf-8")
    except UnicodeDecodeError:
        text = _read("latin-1")

    logger.info("CSV parsed: %d chars", len(text))
    return text


def _clean_text(text: str) -> str:
    """Remove common header/footer artifacts and normalize whitespace."""
    # Remove page numbers like "- 1 -" or "Sayfa 1/10"
    text = re.sub(r"[-–]\s*\d+\s*[-–]", " ", text)
    text = re.sub(r"[Ss]ayfa\s+\d+\s*/\s*\d+", " ", text)
    text = re.sub(r"\bPage\s+\d+\b", " ", text, flags=re.IGNORECASE)

    # Collapse excessive blank lines (more than 2 in a row)
    text = re.sub(r"\n{3,}", "\n\n", text)

    # Normalize unicode whitespace characters
    text = re.sub(r"[\t\r\xa0​]", " ", text)

    # Collapse multiple spaces
    text = re.sub(r" {2,}", " ", text)

    return text.strip()


if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    if len(sys.argv) < 2:
        print("Usage: python -m ingestion.parser <file_path>")
        sys.exit(1)
    result = parse_document(sys.argv[1])
    print(result[:2000])
    print(f"\n[Total: {len(result)} characters]")
