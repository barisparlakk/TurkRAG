"""Tests for document parser — TXT and CSV (no binary deps required)."""

import csv
import importlib

import pytest

from ingestion import parser
from ingestion.parser import parse_document


class TestTxtParsing:
    def test_basic_txt(self, tmp_path):
        f = tmp_path / "doc.txt"
        f.write_text("Merhaba dünya.\nBu bir test.", encoding="utf-8")
        text = parse_document(str(f))
        assert "Merhaba" in text
        assert "test" in text

    def test_unsupported_extension_raises(self, tmp_path):
        f = tmp_path / "file.xyz"
        f.write_bytes(b"data")
        with pytest.raises(ValueError, match="Unsupported"):
            parse_document(str(f))

    def test_returns_string(self, tmp_path):
        f = tmp_path / "doc.txt"
        f.write_text("metin", encoding="utf-8")
        assert isinstance(parse_document(str(f)), str)

    def test_strips_page_numbers(self, tmp_path):
        f = tmp_path / "doc.txt"
        f.write_text("Paragraf metni.\n- 1 -\nSonraki paragraf.", encoding="utf-8")
        text = parse_document(str(f))
        assert "- 1 -" not in text

    def test_normalizes_whitespace(self, tmp_path):
        f = tmp_path / "doc.txt"
        f.write_text("kelime1   \t  kelime2\n\n\nkelime3", encoding="utf-8")
        text = parse_document(str(f))
        assert "kelime1" in text and "kelime2" in text

    def test_latin1_fallback(self, tmp_path):
        f = tmp_path / "doc.txt"
        f.write_bytes("İçerik burada.".encode("latin-1", errors="replace"))
        text = parse_document(str(f))
        assert isinstance(text, str)


class TestCsvParsing:
    def _write_csv(self, tmp_path, rows, filename="data.csv"):
        f = tmp_path / filename
        with open(f, "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
        return str(f)

    def test_basic_csv(self, tmp_path):
        path = self._write_csv(tmp_path, [
            {"İsim": "Ahmet", "Departman": "IT"},
            {"İsim": "Ayşe", "Departman": "İK"},
        ])
        text = parse_document(path)
        assert "Ahmet" in text
        assert "İsim" in text

    def test_csv_key_value_format(self, tmp_path):
        path = self._write_csv(tmp_path, [{"Ad": "Test", "Soyad": "Kişi"}])
        text = parse_document(path)
        assert "Ad: Test" in text

    def test_empty_cells_omitted(self, tmp_path):
        path = self._write_csv(tmp_path, [{"Ad": "Ali", "Soyad": ""}])
        text = parse_document(path)
        assert "Soyad" not in text

    def test_multiple_rows(self, tmp_path):
        path = self._write_csv(tmp_path, [
            {"Ürün": f"ürün{i}", "Fiyat": str(i * 10)}
            for i in range(5)
        ])
        text = parse_document(path)
        for i in range(5):
            assert f"ürün{i}" in text

    def test_csv_returns_nonempty(self, tmp_path):
        path = self._write_csv(tmp_path, [{"A": "1", "B": "2"}])
        assert len(parse_document(path)) > 0

    def test_csv_row_limit_rejects_bomb(self, tmp_path, monkeypatch):
        monkeypatch.setattr(parser, "MAX_SPREADSHEET_ROWS", 1)
        path = self._write_csv(tmp_path, [{"A": "1"}, {"A": "2"}])
        with pytest.raises(ValueError, match="row count"):
            parse_document(path)


class TestExcelParsing:
    def test_basic_xlsx(self, tmp_path):
        pytest.importorskip("openpyxl")
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapor"
        ws.append(["Ad", "Departman"])
        ws.append(["Mehmet", "Mühendislik"])
        path = str(tmp_path / "data.xlsx")
        wb.save(path)

        text = parse_document(path)
        assert "Mehmet" in text
        assert "Mühendislik" in text

    def test_xlsx_sheet_header_present(self, tmp_path):
        pytest.importorskip("openpyxl")
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSayfa"
        ws.append(["Sütun"])
        ws.append(["Değer"])
        path = str(tmp_path / "test.xlsx")
        wb.save(path)

        text = parse_document(path)
        assert "TestSayfa" in text

    def test_xlsx_key_value_format(self, tmp_path):
        pytest.importorskip("openpyxl")
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["İsim", "Görev"])
        ws.append(["Ali", "Müdür"])
        path = str(tmp_path / "kv.xlsx")
        wb.save(path)

        text = parse_document(path)
        assert "İsim: Ali" in text
        assert "Görev: Müdür" in text

    def test_empty_xlsx_returns_empty_or_minimal(self, tmp_path):
        pytest.importorskip("openpyxl")
        import openpyxl

        wb = openpyxl.Workbook()
        path = str(tmp_path / "empty.xlsx")
        wb.save(path)
        text = parse_document(path)
        assert isinstance(text, str)

    def test_xlsx_cell_limit_rejects_bomb(self, tmp_path, monkeypatch):
        pytest.importorskip("openpyxl")
        import openpyxl

        monkeypatch.setattr(parser, "MAX_SPREADSHEET_CELLS", 1)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        path = str(tmp_path / "wide.xlsx")
        wb.save(path)

        with pytest.raises(ValueError, match="size limits"):
            parse_document(path)


@pytest.mark.parametrize(
    ("env_name", "attr_name", "valid_value"),
    [
        ("MAX_PARSED_CHARS", "MAX_PARSED_CHARS", "4096"),
        ("MAX_PDF_PAGES", "MAX_PDF_PAGES", "12"),
        ("MAX_SPREADSHEET_ROWS", "MAX_SPREADSHEET_ROWS", "100"),
        ("MAX_SPREADSHEET_CELLS", "MAX_SPREADSHEET_CELLS", "500"),
        ("MAX_CSV_FIELD_SIZE", "MAX_CSV_FIELD_SIZE", "2048"),
    ],
)
def test_parser_numeric_env_limits_fail_fast(monkeypatch, env_name, attr_name, valid_value):
    monkeypatch.setenv(env_name, "0")
    with pytest.raises(RuntimeError, match=f"{env_name} must be a positive integer"):
        importlib.reload(parser)

    monkeypatch.setenv(env_name, valid_value)
    reloaded = importlib.reload(parser)
    assert getattr(reloaded, attr_name) == int(valid_value)
